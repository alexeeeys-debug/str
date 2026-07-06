import io
import numpy as np
import pandas as pd
import streamlit as st
import altair as alt
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from anomaly_engine import analyze_clients, load_registry

try:
    import ruptures as rpt
    HAS_RUPTURES = True
except Exception:
    HAS_RUPTURES = False

st.set_page_config(page_title="Контроль начислений", layout="wide")
alt.data_transformers.disable_max_rows()  # страницы по когортам могут содержать много точек

# --- Идентификаторы методов детекции точки перегиба ---
METHOD_SLOPES = "Макс. разница наклонов (текущий)"
METHOD_CUSUM = "CUSUM (смена режима)"
METHOD_SUSTAINED = "Устойчивый выход за уровень"
METHOD_RUPTURES = "ruptures / PELT"
# METHODS = [METHOD_SLOPES, METHOD_CUSUM, METHOD_SUSTAINED, METHOD_RUPTURES]
METHODS = [METHOD_SUSTAINED]


def get_severity_color(level, severity):
    lvl = str(level).lower()
    if 'высок' in lvl or 'критич' in lvl:
        return 'FF0000', True
    elif 'средн' in lvl:
        return 'FFA500', False
    elif 'низк' in lvl:
        return 'FFFF00', False

    if severity >= 0.7:
        return 'FF0000', True
    elif severity >= 0.4:
        return 'FFA500', False
    else:
        return 'FFFF00', False


# ==================================================================
# ДЕТЕКЦИЯ ТОЧКИ ПЕРЕГИБА — общие помощники
# Все детекторы возвращают dict {score, inflection_idx, early_slope, late_slope} либо None.
# ==================================================================
def _prep(series, threshold_ratio=0.01, min_active=6):
    """Чистка ряда, поиск старта, отсечение микро-сумм. -> (start, active) или None."""
    arr = np.nan_to_num(np.array(series, dtype=float))
    non_zero = arr[arr > 0]
    if len(non_zero) < min_active:
        return None
    threshold = np.median(non_zero) * threshold_ratio
    floor = threshold if threshold > 0 else 1e-9
    valid = np.where(arr > threshold)[0]
    if len(valid) == 0:
        return None
    start = valid[0]
    active = np.clip(arr[start:], floor, None)
    if len(active) < min_active:
        return None
    return start, active


def _median_slopes(active, k):
    """Робастные наклоны (медиана лог-приростов) слева и справа от точки k (в active)."""
    g = np.diff(np.log(active))
    early = float(np.median(g[:k])) if k >= 1 else 0.0
    late = float(np.median(g[k:])) if k < len(g) else 0.0
    return early, late


def _seg_slope(a, b, Sy, Siy):
    """Наклон линейной регрессии для сегмента [a, b) по префикс-суммам (O(1))."""
    m = b - a
    if m < 2:
        return 0.0
    sum_y = Sy[b] - Sy[a]
    sum_iy = (Siy[b] - Siy[a]) - a * sum_y
    sum_x = m * (m - 1) / 2.0
    sum_x2 = (m - 1) * m * (2 * m - 1) / 6.0
    denom = m * sum_x2 - sum_x * sum_x
    if denom == 0:
        return 0.0
    return (m * sum_iy - sum_x * sum_y) / denom


# ---- Метод 1: максимальная разница наклонов OLS (исходный) ----
def detect_breakout_slopes(series, threshold_ratio=0.01, min_active=6):
    prep = _prep(series, threshold_ratio, min_active)
    if prep is None:
        return None
    start, active = prep
    n = len(active)

    log_active = np.log(active)
    idx = np.arange(n, dtype=float)
    Sy = np.concatenate([[0.0], np.cumsum(log_active)])
    Siy = np.concatenate([[0.0], np.cumsum(idx * log_active)])

    best = None
    for k in range(2, n - 1):
        s_early = _seg_slope(0, k, Sy, Siy)
        s_late = _seg_slope(k, n, Sy, Siy)
        accel = s_late - s_early
        if best is None or accel > best['score']:
            best = {'score': float(accel), 'inflection_idx': int(start + k),
                    'early_slope': float(s_early), 'late_slope': float(s_late)}
    return best


# ---- Метод 2: CUSUM по лог-приростам (смена режима роста) ----
def detect_breakout_cusum(series, threshold_ratio=0.01, min_active=6, slack=0.10, h=0.60):
    prep = _prep(series, threshold_ratio, min_active)
    if prep is None:
        return None
    start, active = prep
    n = len(active)

    g = np.diff(np.log(active))
    mu = np.median(g)

    S = 0.0
    reset = 0
    peakS, peak_reset = -1.0, 0
    for i in range(len(g)):
        S = max(0.0, S + (g[i] - mu - slack))
        if S == 0.0:
            reset = i + 1
        if S > peakS:
            peakS, peak_reset = S, reset

    if peakS < h:
        return None
    k = peak_reset
    if k < 1 or k >= n:
        return None

    early, late = _median_slopes(active, k)
    return {'score': float(late - early), 'inflection_idx': int(start + k),
            'early_slope': early, 'late_slope': late}


# ---- Метод 3: устойчивый выход за базовый уровень (интерпретируемый) ----
def detect_breakout_sustained(series, threshold_ratio=0.01, min_active=6,
                              level_mult=2.0, min_run=2, share=0.6, smooth=3):
    prep = _prep(series, threshold_ratio, min_active)
    if prep is None:
        return None
    start, active = prep
    n = len(active)

    sm = pd.Series(active).rolling(smooth, min_periods=1, center=True).median().to_numpy()
    base_win = max(4, n // 3)
    base = np.median(active[:base_win])
    level = base * level_mult

    k = None
    for i in range(1, n - int(min_run) + 1):
        run_ok = np.all(sm[i:i + int(min_run)] >= level)
        tail_ok = np.mean(sm[i:] >= level) >= share
        if run_ok and tail_ok:
            k = i
            break
    if k is None:
        return None

    early, late = _median_slopes(active, k)
    return {'score': float(late - early), 'inflection_idx': int(start + k),
            'early_slope': early, 'late_slope': late}


# ---- Метод 4: ruptures / PELT (промышленный changepoint) ----
def detect_breakout_ruptures(series, threshold_ratio=0.01, min_active=6, pen=0.5):
    if not HAS_RUPTURES:
        return None
    prep = _prep(series, threshold_ratio, min_active)
    if prep is None:
        return None
    start, active = prep

    g = np.diff(np.log(active))
    if len(g) < 4:
        return None
    try:
        algo = rpt.Pelt(model="l2", min_size=2).fit(g.reshape(-1, 1))
        bkps = algo.predict(pen=float(pen))
    except Exception:
        return None

    candidates = [b for b in bkps if 1 <= b < len(g)]
    if not candidates:
        return None

    best = None
    for bk in candidates:
        early, late = _median_slopes(active, bk)
        sc = late - early
        if best is None or sc > best['score']:
            best = {'score': float(sc), 'inflection_idx': int(start + bk),
                    'early_slope': float(early), 'late_slope': float(late)}
    return best


def _detect(series, method, slack, h, level_mult, min_run, pen):
    """Диспетчер: выбирает детектор по имени метода."""
    if method == METHOD_CUSUM:
        return detect_breakout_cusum(series, slack=slack, h=h)
    if method == METHOD_SUSTAINED:
        return detect_breakout_sustained(series, level_mult=level_mult, min_run=min_run)
    if method == METHOD_RUPTURES:
        return detect_breakout_ruptures(series, pen=pen)
    return detect_breakout_slopes(series)


@st.cache_data
def calculate_market_and_cohorts(clients, quarters, method, breakout_thr, breakout_min_slope,
                                 slack, h, level_mult, min_run, pen):
    """
    Расчёт глобального рынка, разбивка на когорты и пометка 🔥 взрывного роста
    выбранным методом детекции точки перегиба.
    """
    q_list = list(quarters)
    market_sum = np.zeros(len(q_list))
    active_counts = np.zeros(len(q_list))

    for c in clients:
        arr = np.nan_to_num(c['series'])
        market_sum += arr
        active_counts += (arr > 0).astype(int)

    active_counts = np.where(active_counts == 0, 1, active_counts)
    market_series = market_sum / active_counts

    client_data = []

    for c in clients:
        arr = np.nan_to_num(c['series'])

        check_len = min(3, len(arr))
        if not np.all(arr[-check_len:] > 0):
            continue

        non_zero = arr[arr > 0]
        threshold = np.median(non_zero) * 0.01 if len(non_zero) > 0 else 0
        valid_idx = np.where(arr > threshold)[0]
        start_idx = valid_idx[0] if len(valid_idx) > 0 else 0

        active_c_series = arr[start_idx:]
        active_m_series = market_series[start_idx:]

        if len(active_c_series) < 3:
            c_base = np.median(active_c_series) if np.median(active_c_series) != 0 else 1
            m_base = np.median(active_m_series) if np.median(active_m_series) != 0 else 1
            c_last = active_c_series[-1]
            m_last = active_m_series[-1]
        else:
            first_3_client = active_c_series[active_c_series > threshold][:3]
            c_base = np.median(first_3_client) if (len(first_3_client) > 0 and np.median(first_3_client) != 0) else 1
            first_3_market = active_m_series[:3]
            m_base = np.median(first_3_market) if (len(first_3_market) > 0 and np.median(first_3_market) != 0) else 1
            c_last = np.median(active_c_series[-3:])
            m_last = np.median(active_m_series[-3:])

        cl_growth = (c_last / c_base) - 1
        mkt_growth_for_client = (m_last / m_base) - 1
        delta = cl_growth - mkt_growth_for_client

        if delta > 0.15:
            cohort = "🚀 Выше рынка"
        elif delta < -0.15:
            cohort = "📉 Ниже рынка"
        else:
            cohort = "🤝 В рынке"

        # --- ТОЧКА ПЕРЕГИБА выбранным методом ---
        bo = _detect(arr, method, slack, h, level_mult, min_run, pen)
        accel = float(bo['score']) if bo else 0.0
        late_rate = float(np.expm1(bo['late_slope'])) if bo else np.nan
        infl_q = q_list[bo['inflection_idx']] if (bo and bo['inflection_idx'] < len(q_list)) else None

        is_breakout = (cohort == "🚀 Выше рынка"
                       and bo is not None
                       and accel >= breakout_thr
                       and bo['late_slope'] >= breakout_min_slope)

        subtype = "🔥 Взрывной рост" if is_breakout else cohort

        client_data.append({'ID': str(c['id']), 'Тип': c.get('ctype', ''), 'Когорта': cohort,
                            'Подтип': subtype, 'Взрывной рост': bool(is_breakout),
                            'Ускорение': accel, 'Темп посл.': late_rate, 'Квартал перегиба': infl_q,
                            'Рост клиента': cl_growth, 'Отрыв от рынка': delta, 'series': arr})

    df_clients = pd.DataFrame(client_data)
    return q_list, market_series, df_clients


def _normalize_client(cser, market_series, q_list):
    """Нормализация клиента и рынка к базе=100 на индивидуальном старте клиента.
    Возвращает (active_q_list, mnorm, cnorm) либо None, если истории мало."""
    cser = np.nan_to_num(np.array(cser, dtype=float))
    nz = cser[cser > 0]
    thr = np.median(nz) * 0.01 if len(nz) > 0 else 0
    vidx = np.where(cser > thr)[0]
    s0 = vidx[0] if len(vidx) > 0 else 0
    aq = q_list[s0:]
    am = market_series[s0:]
    ac = cser[s0:]
    if len(aq) < 2:
        return None
    f3c = ac[ac > thr][:3]
    cbase = np.median(f3c) if (len(f3c) > 0 and np.median(f3c) != 0) else 1
    f3m = am[:3]
    mbase = np.median(f3m) if (len(f3m) > 0 and np.median(f3m) != 0) else 1
    return aq, (am / mbase) * 100.0, (ac / cbase) * 100.0


@st.cache_data
def build_cohort_longdf(clients, quarters, cohort_label, method, breakout_thr, breakout_min_slope,
                        slack, h, level_mult, min_run, pen):
    """
    Готовит «длинный» датафрейм для фасетной сетки графиков по когорте.
    Клиенты отсортированы по ID по убыванию; для каждого — линии «Рынок» и «Клиент»
    (индекс роста, база=100) и отметка квартала перегиба (для 🔥).
    Кэшируется по методу/параметрам — перестройка при неизменных входах бесплатна.
    """
    q_list, market_series, df_clients = calculate_market_and_cohorts(
        clients, quarters, method, breakout_thr, breakout_min_slope,
        slack, h, level_mult, min_run, pen)

    if df_clients.empty:
        return q_list, [], pd.DataFrame(columns=['ID', 'Период', 'Значение', 'Показатель', 'is_infl'])

    if cohort_label == "🔥 Взрывной рост":
        sub = df_clients[df_clients['Взрывной рост']]
    else:
        sub = df_clients[df_clients['Когорта'] == cohort_label]

    def _idnum(v):
        try:
            return int(v)
        except Exception:
            return float('-inf')

    ordered_ids = sorted(sub['ID'].tolist(), key=_idnum, reverse=True)
    by_id = {r['ID']: r for _, r in sub.iterrows()}

    rows = []
    for cid in ordered_ids:
        row = by_id[cid]
        norm = _normalize_client(row['series'], market_series, q_list)
        if norm is None:
            continue
        aq, mnorm, cnorm = norm
        infl = row['Квартал перегиба'] if bool(row['Взрывной рост']) else None
        for i, q in enumerate(aq):
            rows.append({'ID': cid, 'Период': q, 'Значение': float(mnorm[i]),
                         'Показатель': 'Рынок', 'is_infl': False})
            rows.append({'ID': cid, 'Период': q, 'Значение': float(cnorm[i]),
                         'Показатель': 'Клиент', 'is_infl': bool(infl is not None and q == infl)})

    df_long = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['ID', 'Период', 'Значение', 'Показатель', 'is_infl'])
    # оставляем только реально отрисованных клиентов, порядок сохраняем
    drawn = [cid for cid in ordered_ids if cid in set(df_long['ID'])] if not df_long.empty else []
    return q_list, drawn, df_long


def build_annotated(wb, ws, qcol, clients, flags):
    from openpyxl.utils import get_column_letter

    row_of = {str(c['id']): c['row'] for c in clients}
    for f in flags:
        cell = ws.cell(row_of[f['client_id']], qcol[f['quarter']])

        hexc, dark = get_severity_color(f['level'], f['severity'])

        cell.fill = PatternFill('solid', fgColor=hexc)
        if dark:
            cell.font = Font(name=cell.font.name or 'Arial', color='FFFFFF', bold=True)
        exp = f"\nОбычный диапазон: {f['lo']:,.0f} … {f['hi']:,.0f}" if f.get('lo') else ''
        arrow = 'ВЫШЕ ожидаемого ↑' if f['direction'] > 0 else 'НИЖЕ ожидаемого ↓'
        cm = Comment(f"⚠ {arrow} | {f['level']}\nЗначение: {f['value']:,.2f}\n\n" +
                     "\n".join('• ' + r for r in f['reasons']) + exp, "Контроль")
        cm.width = 320;
        cm.height = 240;
        cell.comment = cm

    if clients:
        header_row = min(c['row'] for c in clients) - 1
        max_col_letter = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A{header_row}:{max_col_letter}{ws.max_row}"

    bio = io.BytesIO();
    wb.save(bio);
    return bio.getvalue()


# Цветовая схема подтипов когорт (единая для круговой диаграммы и карты распределения)
SUBTYPE_DOMAIN = ["🔥 Взрывной рост", "🚀 Выше рынка", "🤝 В рынке", "📉 Ниже рынка"]
# SUBTYPE_RANGE = ["#E67E22", "#2ECC71", "#F1C40F", "#E74C3C"]
SUBTYPE_RANGE = ["#196F3D", "#2ECC71", "#F1C40F", "#E74C3C"]


def render_cohort_facets(cohort_label, clients, quarters, method, breakout_thr, breakout_min_slope,
                         slack, h, level_mult, min_run, pen):
    """Страница когорты: сетка малых графиков «клиент vs рынок» по каждому клиенту."""
    q_list, ordered_ids, df_long = build_cohort_longdf(
        clients, quarters, cohort_label, method, breakout_thr, breakout_min_slope,
        slack, h, level_mult, int(min_run), pen)

    n = len(ordered_ids)
    st.subheader(f"{cohort_label} — {n} клиентов")
    if n == 0 or df_long.empty:
        st.info("В этой когорте нет клиентов.")
        return

    st.caption("Сортировка по ID (от большего к меньшему). Каждая панель — индекс роста (база=100): "
               "оранжевым клиент, серым рынок. Красная точка — квартал перегиба (для 🔥). "
               "Наведите курсор для значений.")
    cols = st.radio("Колонок в сетке", [2, 3, 4, 5], index=2, horizontal=True,
                    key=f"cols_{cohort_label}")

    enc_x = alt.X('Период:N', sort=q_list, axis=alt.Axis(labels=False, ticks=False, title=None))
    lines = alt.Chart(df_long).mark_line(strokeWidth=2).encode(
        x=enc_x,
        y=alt.Y('Значение:Q', title=None),
        color=alt.Color('Показатель:N',
                        scale=alt.Scale(domain=['Рынок', 'Клиент'], range=['#95A5A6', '#E67E22']),
                        legend=alt.Legend(orient='top', title=None)),
        tooltip=['ID', 'Период', 'Показатель', alt.Tooltip('Значение:Q', format=',.0f')])
    pts = alt.Chart(df_long).mark_point(size=65, filled=True, color='#C0392B').encode(
        x=enc_x, y=alt.Y('Значение:Q'),
        tooltip=['ID', 'Период', alt.Tooltip('Значение:Q', format=',.0f')]
    ).transform_filter(alt.datum.is_infl == True)

    grid = alt.layer(lines, pts).properties(width=190, height=120).facet(
        facet=alt.Facet('ID:N', sort=ordered_ids, title=None,
                        header=alt.Header(labelFontWeight='bold', labelFontSize=12)),
        columns=int(cols)
    ).resolve_scale(y='independent')

    st.altair_chart(grid, use_container_width=False)


def main():
    st.title("Контроль ошибок заполнения сумм начислений")
    st.caption(
        "Цветовая разметка (независимо от знака отклонения): Жёлтый — низкая важность, Оранжевый — средняя, Красный — высокая. ")

    with st.sidebar:
        st.header("Чувствительность")
        with st.popover("ℹ️ Инструкция по настройкам"):
            st.markdown("""
                    **Основные статистические параметры:**
                    * **Порог робастного z-score:** Насколько сильно текущее значение выбивается из исторического коридора клиента. 3.5 считается оптимальной границей. Чем *ниже* порог, тем *выше* чувствительность.
                    * **Порог скачка (в разах):** Прямолинейный фильтр для поиска грубых ошибок масштаба.
                    * **Мин. история клиента (кварталы):** Ограничение по объему данных, необходимых для достоверного анализа.

                    **Проверка по медиане последних кварталов:**
                    * Полезна для поиска локальных аномалий.
                    * **Окно:** За сколько последних кварталов считать медиану.
                    * **Порог отклонения:** Допустимый процент колебаний.

                    **Взрывной рост (точка перегиба):**
                    * Метод выбирает, *куда* ставится точка перегиба. Ниже — короткое описание каждого.
                    * Общий порог отнесения к 🔥 одинаков для всех методов, чтобы их можно было честно сравнивать.
                    """)
        z_thr = st.slider("Порог робастного z-score", 2.5, 6.0, 3.5, 0.5)
        ratio_thr = st.slider("Порог скачка (в разах)", 3, 20, 8, 1)
        min_history = st.slider("Мин. история клиента (кварталы)", 3, 12, 6, 1)
        st.divider()
        enable_trailing = st.checkbox("Проверка по медиане последних кварталов", value=True)
        trailing_window = st.number_input("Окно (последних кварталов)", min_value=1, max_value=12, value=4, step=1,
                                          disabled=not enable_trailing)
        pct_threshold = st.number_input("Порог отклонения от медианы, %", min_value=1.0, max_value=500.0, value=50.0,
                                        step=5.0, disabled=not enable_trailing)

        st.divider()
        st.subheader("Взрывной рост")
        method = st.selectbox(
            "Метод точки перегиба", METHODS,
            help="Все методы возвращают квартал перегиба; различаются устойчивостью к шуму и тем, "
                 "что считать «переломом» — локальное ускорение или устойчивую смену режима.")

        # Значения по умолчанию (используются, если метод их не запрашивает)
        slack, h, level_mult, min_run, pen = 0.10, 0.60, 2.0, 2, 0.5

        if method == METHOD_SLOPES:
            st.caption("Ищет max разницу наклонов OLS. Чувствителен к ранним всплескам.")
        elif method == METHOD_CUSUM:
            st.caption("Копит устойчивое превышение темпа роста. Одиночные всплески гасятся.")
            slack = st.slider("CUSUM: slack — «шум», лог/кв", 0.0, 0.5, 0.10, 0.02,
                              help="Какой прирост считать нормальным шумом. Выше → устойчивее к всплескам.")
            h = st.slider("CUSUM: порог h", 0.2, 2.0, 0.60, 0.05,
                          help="Минимальная высота накопления, ниже которой сдвига нет.")
        elif method == METHOD_SUSTAINED:
            st.caption("Момент, после которого клиент устойчиво держится выше своего базового уровня.")
            level_mult = st.slider("Множитель уровня (×база)", 1.2, 5.0, 2.0, 0.1,
                                   help="Во сколько раз выше базы считать «новым уровнем».")
            min_run = st.slider("Кварталов подряд выше уровня", 1, 6, 2, 1)
        elif method == METHOD_RUPTURES:
            if not HAS_RUPTURES:
                st.warning("Библиотека не установлена. Выполните: pip install ruptures")
            st.caption("PELT по лог-приростам. Штраф регулирует чувствительность к числу точек.")
            pen = st.slider("ruptures: штраф (pen)", 0.1, 10.0, 0.5, 0.1,
                            help="Выше штраф → меньше точек разбиения (только сильные сдвиги). "
                                 "Для квартальных рядов обычно 0.3–1.0.")

        with st.expander("Порог отнесения к 🔥 (общий для всех методов)"):
            breakout_thr = st.slider("Мин. ускорение вокруг точки (лог/кв)", 0.05, 1.00, 0.20, 0.05,
                                     help="Ниже порог → в 🔥 попадает больше клиентов.")
            breakout_min_slope = st.slider("Мин. темп в новом режиме (лог/кв)", 0.0, 0.5, 0.05, 0.01,
                                           help="Новый режим должен реально расти.")

    up = st.file_uploader("Файл реестра (.xlsx)", type=['xlsx'])
    status_placeholder = st.empty()

    if not up:
        status_placeholder.info("Ожидается загрузка файла.")
        st.stop()

    status_placeholder.success("Выполняется обработка файла...")

    wb = openpyxl.load_workbook(up)
    ws, hdr, col, qcol, quarters, clients = load_registry(wb)

    with st.spinner(f"Анализ {len(clients)} клиентов…"):
        flags = analyze_clients(clients, min_history, z_thr, ratio_thr, z_thr, quarters,
                                trailing_window=int(trailing_window), pct_threshold=float(pct_threshold),
                                enable_trailing=bool(enable_trailing))

    status_placeholder.success("Готово")
    st.divider()

    # --- Единый расчёт когорт (кэш) — используется всеми вкладками ---
    q_list, market_series, df_clients = calculate_market_and_cohorts(
        clients, quarters, method, breakout_thr, breakout_min_slope,
        slack, h, level_mult, int(min_run), pen)

    tab_search, tab_market, tab_fire, tab_up, tab_mid, tab_down = st.tabs(
        ["🔍 Поиск ошибок", "📈 Аналитика рынка",
         "🔥 Взрывной рост", "🚀 Выше рынка", "🤝 В рынке", "📉 Ниже рынка"])

    # ==========================================
    # ВКЛАДКА 1: ПОИСК ОШИБОК
    # ==========================================
    with tab_search:
        period_options = ["Все периоды"] + list(quarters)[::-1]
        selected_period = st.selectbox("📅 Выберите период для вывода отчета на экран:", period_options, index=1)

        if selected_period == "Все периоды":
            display_flags = flags
        else:
            display_flags = [f for f in flags if f['quarter'] == selected_period]

        n_up = sum(f['direction'] > 0 for f in display_flags)
        n_dn = sum(f['direction'] < 0 for f in display_flags)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Клиентов (всего)", len(clients))
        c2.metric("Подозрительных", len(display_flags))
        c3.metric("Вверх (↑)", n_up)
        c4.metric("Вниз (↓)", n_dn)

        st.subheader(f"Ранжированный список ({selected_period})")
        rows = []
        for f in display_flags:
            rows.append({'ID клиента': f['client_id'], 'Тип': f['ctype'], 'Резидент': f['resident'],
                         'Период': f['quarter'], 'Значение': round(f['value'], 2),
                         'Напр.': '↑ выше' if f['direction'] > 0 else '↓ ниже',
                         'Ожид. диапазон': (f"{f['lo']:,.0f} … {f['hi']:,.0f}" if f.get('lo') else '—'),
                         'Серьёзность': f['level'], '_sev': f['severity'], '_dir': f['direction'],
                         'Причина(ы)': '; '.join(f['reasons'])})
        df = pd.DataFrame(rows)

        if df.empty:
            st.success(f"🎉 В периоде «{selected_period}» подозрительных начислений не найдено!")
        else:
            def row_style(row):
                hexc, dark = get_severity_color(row['Серьёзность'], row['_sev'])
                style = f'background-color:#{hexc}' + (';color:white' if dark else ';color:black')
                return [style if col in ('Значение', 'Напр.', 'Серьёзность') else '' for col in df.columns]

            st.dataframe(df.style.apply(row_style, axis=1).hide(['_sev', '_dir'], axis=1),
                         use_container_width=True, height=430)

        st.divider()

        dl_col1, dl_col2 = st.columns(2)
        # with dl_col1:
        #     st.download_button(
        #         label="⬇ Скачать размеченный Excel (все периоды)",
        #         data=build_annotated(wb, ws, qcol, clients, flags),
        #         file_name="реестр_начислений_все.xlsx",
        #         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        #     )

        with dl_col1:
            wb_filtered = openpyxl.load_workbook(up)
            ws_filtered = wb_filtered[ws.title]
            st.download_button(
                label=f"⬇ Скачать Excel (размечен {selected_period})",
                data=build_annotated(wb_filtered, ws_filtered, qcol, clients, display_flags),
                file_name=f"реестр_начислений_{selected_period}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                disabled=(selected_period == "Все периоды")
            )

        st.subheader("Профиль клиента (детализация аномалии)")
        ids = [str(c['id']) for c in clients]
        default_ix = ids.index(display_flags[0]['client_id']) if display_flags else (
            ids.index(flags[0]['client_id']) if flags else 0)

        sel = st.selectbox("Клиент", ids, index=default_ix)
        c = next(c for c in clients if str(c['id']) == sel)
        fmap = {f['quarter']: f['direction'] for f in flags if f['client_id'] == sel}
        cd = pd.DataFrame({'Период': quarters, 'Сумма': c['series'],
                           'Статус': ['выше ↑' if fmap.get(q, 0) > 0 else 'ниже ↓' if fmap.get(q, 0) < 0 else 'норма'
                                      for q in quarters]}).dropna(subset=['Сумма'])
        base = alt.Chart(cd).encode(x=alt.X('Период:N', sort=quarters))
        line = base.mark_line(color='#95A5A6').encode(y=alt.Y('Сумма:Q', scale=alt.Scale(type='log')))
        pts = base.mark_circle(size=95).encode(
            y='Сумма:Q',
            color=alt.Color('Статус:N', scale=alt.Scale(domain=['норма', 'выше ↑', 'ниже ↓'],
                                                        range=['#3498DB', '#C0392B', '#1E8449'])),
            tooltip=['Период', 'Сумма', 'Статус'])
        st.altair_chart((line + pts).properties(height=340), use_container_width=True)

    # ==========================================
    # ВКЛАДКА 2: АНАЛИТИКА РЫНКА
    # ==========================================
    with tab_market:
        st.subheader("Глобальный тренд (в среднем на активного клиента)")
        st.caption(
            "Рыночный индекс рассчитывается не по простой сумме (тоталу), а по **среднему начислению на одного активного клиента** в каждом квартале. Это исключает искажения из-за притока новых или оттока старых клиентов.")

        df_market = pd.DataFrame({'Период': q_list, 'Среднее по рынку': market_series})
        market_chart = alt.Chart(df_market).mark_line(point=True, color='#2E86C1', strokeWidth=3).encode(
            x=alt.X('Период:N', sort=q_list),
            y=alt.Y('Среднее по рынку:Q'),
            tooltip=['Период', alt.Tooltip('Среднее по рынку:Q', format=",.2f")]
        ).properties(height=300)
        st.altair_chart(market_chart, use_container_width=True)

        st.divider()

        n_above = int((df_clients['Когорта'] == "🚀 Выше рынка").sum())
        n_breakout = int(df_clients['Взрывной рост'].sum())
        m1, m2, m3 = st.columns(3)
        m1.metric("Клиентов в анализе", len(df_clients))
        m2.metric("🚀 Выше рынка", n_above)
        m3.metric("🔥 Взрывной рост", n_breakout,
                  help=f"Метод: {method}. Подмножество «выше рынка» с точкой перегиба.")

        st.subheader("Анализ когорт")
        st.caption("Сектор «выше рынка» разделён на 🔥 **взрывной рост** (резкий излом тренда) и остальных стабильно "
                   "растущих клиентов — оба находятся выше рынка.")

        col_c1, col_c2 = st.columns([1, 2])
        with col_c1:
            st.write("**Распределение клиентов:**")
            cohort_counts = df_clients['Подтип'].value_counts().reset_index()
            cohort_counts.columns = ['Подтип', 'Количество']
            pie_chart = alt.Chart(cohort_counts).mark_arc(innerRadius=40).encode(
                theta=alt.Theta(field="Количество", type="quantitative"),
                color=alt.Color(field="Подтип", type="nominal", sort=SUBTYPE_DOMAIN,
                                scale=alt.Scale(domain=SUBTYPE_DOMAIN, range=SUBTYPE_RANGE)),
                order=alt.Order("Подтип:N", sort="ascending"),
                tooltip=['Подтип', 'Количество']
            ).properties(height=250)
            st.altair_chart(pie_chart, use_container_width=True)

        with col_c2:
            st.write("**Карта распределения клиентов:**")
            st.caption("По горизонтали — отрыв от рынка (правее пунктира = «выше рынка»); по вертикали — ускорение "
                       "роста. 🔥 всплывают вверх-вправо, оставаясь внутри зоны «выше рынка».")
            boundary = alt.Chart(pd.DataFrame({'x': [0.15]})).mark_rule(
                color='#7F8C8D', strokeDash=[6, 4]).encode(x='x:Q')
            scatter = alt.Chart(df_clients).mark_circle(opacity=0.75, stroke='white', strokeWidth=0.4).encode(
                x=alt.X('Отрыв от рынка:Q', title='Отрыв от рынка',
                        scale=alt.Scale(type='symlog'), axis=alt.Axis(format='+.0%')),
                y=alt.Y('Ускорение:Q', title='Ускорение роста (перегиб)'),
                size=alt.Size('Взрывной рост:N', scale=alt.Scale(domain=[True, False], range=[240, 55]),
                              legend=None),
                color=alt.Color('Подтип:N', sort=SUBTYPE_DOMAIN,
                                scale=alt.Scale(domain=SUBTYPE_DOMAIN, range=SUBTYPE_RANGE),
                                legend=alt.Legend(title="Подтип")),
                tooltip=['ID', 'Тип', 'Когорта',
                         alt.Tooltip('Рост клиента:Q', format='.0%'),
                         alt.Tooltip('Отрыв от рынка:Q', format='+.0%'),
                         alt.Tooltip('Ускорение:Q', format='.2f'),
                         'Квартал перегиба']
            ).properties(height=280)
            st.altair_chart(boundary + scatter, use_container_width=True)

        st.divider()
        st.write("**Списки клиентов по динамике:**")
        selected_cohort = st.selectbox("Выберите когорту для просмотра",
                                       ["🔥 Взрывной рост", "🚀 Выше рынка", "🤝 В рынке", "📉 Ниже рынка"])

        if selected_cohort == "🔥 Взрывной рост":
            df_filtered = df_clients[df_clients['Взрывной рост']].drop(columns=['series'])
            df_filtered = df_filtered.sort_values('Ускорение', ascending=False)
        else:
            df_filtered = df_clients[df_clients['Когорта'] == selected_cohort].drop(columns=['series'])

        show_cols = ['ID', 'Тип', 'Взрывной рост', 'Рост клиента', 'Отрыв от рынка', 'Темп посл.', 'Квартал перегиба']
        df_show = df_filtered[show_cols]
        st.dataframe(
            df_show.style.format({'Рост клиента': "{:.1%}", 'Отрыв от рынка': "{:+.1%}", 'Темп посл.': "{:+.1%}"},
                                 na_rep='—'),
            height=220, use_container_width=True)

        st.divider()
        st.subheader("Сравнение клиента с рынком")
        st.caption(
            "График показывает **Индекс роста**, где объем первого квартала принят за 100%. Это позволяет визуально сравнивать траектории. "
            "Для клиентов со 🔥 взрывным ростом отдельно отмечается **квартал перегиба** (по выбранному методу).")

        if not df_filtered.empty:
            sel_client_cohort = st.selectbox(
                "Выберите клиента из когорты для сравнения",
                df_filtered['ID'].tolist(),
                key="client_cohort_compare_box"
            )

            if sel_client_cohort:
                cl_row = df_clients[df_clients['ID'] == sel_client_cohort].iloc[0]
                cl_series = np.nan_to_num(np.array(cl_row['series'], dtype=float))

                non_zero = cl_series[cl_series > 0]
                threshold = np.median(non_zero) * 0.01 if len(non_zero) > 0 else 0
                valid_idx = np.where(cl_series > threshold)[0]
                start_idx = valid_idx[0] if len(valid_idx) > 0 else 0

                active_q_list = q_list[start_idx:]
                active_m_series = market_series[start_idx:]
                active_c_series = cl_series[start_idx:]

                first_3_client = active_c_series[active_c_series > threshold][:3]
                c_base = np.median(first_3_client) if (
                            len(first_3_client) > 0 and np.median(first_3_client) != 0) else 1
                first_3_market = active_m_series[:3]
                m_base = np.median(first_3_market) if (
                            len(first_3_market) > 0 and np.median(first_3_market) != 0) else 1

                m_norm = (active_m_series / m_base) * 100
                c_norm = (active_c_series / c_base) * 100

                df_comp = pd.DataFrame({
                    'Период': active_q_list * 2,
                    'Значение': np.concatenate([m_norm, c_norm]),
                    'Показатель': ['Рынок'] * len(active_q_list) + [f'Клиент {sel_client_cohort}'] * len(
                        active_q_list)
                })

                comp_chart = alt.Chart(df_comp).mark_line(point=True, strokeWidth=3).encode(
                    x=alt.X('Период:N', sort=active_q_list),
                    y=alt.Y('Значение:Q', title='Индекс (База = 100)'),
                    color=alt.Color('Показатель:N', scale=alt.Scale(domain=['Рынок', f'Клиент {sel_client_cohort}'],
                                                                    range=['#95A5A6', '#E67E22'])),
                    tooltip=['Период', 'Показатель', alt.Tooltip('Значение:Q', format=",.1f")]
                ).properties(height=350)

                infl_q = cl_row.get('Квартал перегиба')
                if bool(cl_row.get('Взрывной рост')) and infl_q in active_q_list:
                    st.caption(f"🔥 Точка перегиба: **{infl_q}** — здесь рост клиента резко ускорился "
                               f"(метод: {method}).")
                    infl_rule = alt.Chart(pd.DataFrame({'Период': [infl_q]})).mark_rule(
                        color='#E67E22', strokeDash=[5, 5], strokeWidth=2).encode(
                        x=alt.X('Период:N', sort=active_q_list))
                    st.altair_chart(infl_rule + comp_chart, use_container_width=True)
                else:
                    st.altair_chart(comp_chart, use_container_width=True)
        else:
            st.info("В выбранной когорте нет клиентов.")

    # ==========================================
    # ВКЛАДКИ 3–6: СТРАНИЦЫ ПО КОГОРТАМ (сетки графиков по каждому клиенту)
    # ==========================================
    with tab_fire:
        render_cohort_facets("🔥 Взрывной рост", clients, quarters, method, breakout_thr,
                             breakout_min_slope, slack, h, level_mult, min_run, pen)
    with tab_up:
        render_cohort_facets("🚀 Выше рынка", clients, quarters, method, breakout_thr,
                             breakout_min_slope, slack, h, level_mult, min_run, pen)
    with tab_mid:
        render_cohort_facets("🤝 В рынке", clients, quarters, method, breakout_thr,
                             breakout_min_slope, slack, h, level_mult, min_run, pen)
    with tab_down:
        render_cohort_facets("📉 Ниже рынка", clients, quarters, method, breakout_thr,
                             breakout_min_slope, slack, h, level_mult, min_run, pen)


if __name__ == '__main__':
    main()